#!/usr/bin/env python3
"""
Kattis Scraper using Playwright (handles anti-bot protection)

Scrapes solved problems from a Kattis profile and generates statistics.
"""

import argparse
import json
import os
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional
from datetime import datetime, timezone

from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, Page, Browser

# -----------------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------------

KATTIS_BASE_URL = "https://open.kattis.com"
SOLUTION_EXTENSIONS = {'.py', '.cpp', '.c', '.java', '.js', '.go', '.rs', '.kt', '.cs', '.rb'}
MAX_PAGINATION_PAGES = 58  # As of 25/1/2026 this is the maximum number of pages


# -----------------------------------------------------------------------------
# Scraping
# -----------------------------------------------------------------------------

def login_to_kattis(page: Page, username: str, password: str) -> bool:
    """
    Log in to Kattis and return True if successful.
    """
    page.goto(f"{KATTIS_BASE_URL}/login")
    page.wait_for_load_state("networkidle")
    
    # Click "Log in with e-mail" button
    email_login_btn = page.locator('text=Log in with e-mail')
    if email_login_btn.count() > 0:
        email_login_btn.click()
        page.wait_for_load_state("networkidle")
    
    # Fill and submit credentials
    page.fill('input[name="user"]', username)
    page.fill('input[name="password"]', password)
    page.click('text=Submit')
    page.wait_for_load_state("networkidle")
    
    # Check if we're still on login page (failed)
    return "login" not in page.url.lower()

def get_user_total_score(page: Page, username: str) -> float:
    """
    Scrape the actual total score from the user's profile page.
    """
    url = f"{KATTIS_BASE_URL}/users/{username}"
    page.goto(url)
    page.wait_for_load_state("networkidle")
    
    html = page.content()
    soup = BeautifulSoup(html, 'html.parser')
    
    # Try to find the score display
    # Common patterns: look for "Score: XXX.X" or a score element
    
    # Method 1: Look for text pattern "Score: XXX"
    text = soup.get_text()
    score_match = re.search(r'Score[:\s]+(\d+\.?\d*)', text, re.I)
    if score_match:
        try:
            return float(score_match.group(1))
        except ValueError:
            pass
    
    # Method 2: Look for specific score elements (update based on actual HTML)
    score_elem = soup.find('span', class_='score')
    if score_elem:
        try:
            return float(score_elem.get_text(strip=True))
        except ValueError:
            pass
    
    # Method 3: Look for large numbers that are likely scores
    for elem in soup.find_all(['span', 'div', 'td', 'strong']):
        text = elem.get_text(strip=True)
        if re.match(r'^\d{2,4}\.\d$', text):  # Pattern like 595.4 or 1234.5
            try:
                score = float(text)
                if 100 <= score <= 10000:  # Reasonable score range
                    return score
            except ValueError:
                pass
    
    return 0.0

def collect_all_pages_html(page: Page, username: str) -> list[dict]:
    """
    Iterate through all pages and extract problems from each page.
    
    Returns a list of all problems found across all pages.
    """
    all_problems = []
    seen_ids = set()
    page_num = 1
    
    while page_num <= MAX_PAGINATION_PAGES:
        url = f"{KATTIS_BASE_URL}/users/{username}?page={page_num}&tab=problems"
        print(f"  Fetching page {page_num}...")
        
        page.goto(url)
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(1000)
        
        html = page.content()
        
        # Parse problems from this page
        page_problems = extract_problems_from_page(html)
        
        if len(page_problems) == 0:
            print(f"  Reached end of data at page {page_num} (no problems found)")
            break
        
        # Add unique problems
        new_count = 0
        for prob in page_problems:
            if prob['id'] not in seen_ids:
                seen_ids.add(prob['id'])
                all_problems.append(prob)
                new_count += 1
        
        print(f"    Found {len(page_problems)} problems on page, {new_count} new unique problems")
        page_num += 1

        if (new_count == 0):
            break
    
    print(f"  Collected {page_num - 1} page(s) total, {len(all_problems)} total unique problems")
    return all_problems

def extract_problems_from_page(html: str) -> list[dict]:
    """
    Parse HTML from a single page and extract solved problems.
    Only looks at the table under "Solved problems" heading.
    """
    soup = BeautifulSoup(html, 'html.parser')
    problems = []
    seen_ids = set()
    
    # Find the "Solved problems" heading
    solved_heading = soup.find(['h2', 'h3'], string=re.compile(r'Solved\s+problems', re.I))
    
    if not solved_heading:
        return problems
    
    # Find the next table after the heading
    solved_table = solved_heading.find_next('table')
    
    if not solved_table:
        return problems
    
    # Extract problems from this table
    problem_links = solved_table.find_all('a', href=re.compile(r'/problems/[a-z0-9]+$', re.I))
    
    for link in problem_links:
        href = link.get('href', '')
        problem_id = href.split('/problems/')[-1].split('/')[0].split('?')[0]
        
        if not problem_id or problem_id in seen_ids:
            continue
        seen_ids.add(problem_id)
        
        problem_name = link.get_text(strip=True)
        
        # Find difficulty in the same row
        difficulty = 0.0
        row = link.find_parent('tr')
        if row:
            diff_span = row.find('span', class_='difficulty_number')
            if diff_span:
                try:
                    difficulty_text = diff_span.get_text(strip=True)
                    # Handle ranges like "1.2 - 4.1" by taking the value after the hyphen
                    if ' - ' in difficulty_text:
                        difficulty = float(difficulty_text.split(' - ')[1])
                    else:
                        difficulty = float(difficulty_text)
                except (ValueError, IndexError):
                    pass
        
        problems.append({
            'id': problem_id,
            'name': problem_name,
            'difficulty': difficulty,
            'url': f"{KATTIS_BASE_URL}/problems/{problem_id}"
        })
    
    return problems

def scrape_kattis(
    username: str,
    password: str,
    headless: bool = True,
    debug: bool = False
) -> tuple[list[dict], float]:  # Return both problems and score
    """
    Login to Kattis and scrape all solved problems.
    Returns: (problems, total_score)
    """
    problems = []
    total_score = 0.0
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context()
        page = context.new_page()
        
        # Login
        print("Logging in to Kattis...")
        if not login_to_kattis(page, username, password):
            browser.close()
            raise Exception("Login failed. Check your credentials.")
        print(f"  Logged in successfully")
        
        # Get the actual total score from profile
        print(f"Fetching total score from profile...")
        total_score = get_user_total_score(page, username)
        print(f"  Kattis shows total score: {total_score}")
        
        # Navigate to profile
        print(f"Fetching profile for '{username}'...")
        page.goto(f"{KATTIS_BASE_URL}/users/{username}")
        page.wait_for_load_state("networkidle")
        
        # Collect problems from all pagination pages
        print("Collecting solved problems...")
        problems = collect_all_pages_html(page, username)
        
        # Add debug output
        if debug:
            debug_score_calculation(problems)
            print(f"\nCalculated score from problems: {sum(p['difficulty'] for p in problems):.1f}")
            print(f"Actual score from Kattis: {total_score:.1f}")
        
        browser.close()
    
    return problems, total_score

def debug_score_calculation(problems: list[dict]) -> None:
    """
    Debug the score calculation by showing sample problems and their difficulties.
    """
    print("\n=== Score Debug ===")
    print(f"Total problems: {len(problems)}")
    
    # Show problems with missing difficulties
    no_difficulty = [p for p in problems if p['difficulty'] == 0.0]
    print(f"Problems with 0.0 difficulty: {len(no_difficulty)}")
    if len(no_difficulty) > 0:
        print("Problems without difficulty:")
        for p in no_difficulty:
            print(f"  - {p['id']}: {p['name']} ({p['url']})")
        
        # Save to file for easy access
        with open('zero_difficulty_problems.txt', 'w') as f:
            for p in no_difficulty:
                f.write(f"{p['url']}\n")
        print(f"\nURLs saved to: zero_difficulty_problems.txt")
    
    # Show distribution
    difficulties = [p['difficulty'] for p in problems]
    total_score = sum(difficulties)
    avg_difficulty = total_score / len(problems) if problems else 0
    
    print(f"\nTotal score: {total_score:.1f}")
    print(f"Average difficulty: {avg_difficulty:.2f}")
    print(f"Min difficulty: {min(difficulties) if difficulties else 0:.1f}")
    print(f"Max difficulty: {max(difficulties) if difficulties else 0:.1f}")
    
    # Show sample of highest difficulty problems
    sorted_problems = sorted(problems, key=lambda x: x['difficulty'], reverse=True)
    print(f"\nTop 5 hardest problems:")
    for p in sorted_problems[:5]:
        print(f"  {p['difficulty']:.1f} - {p['id']}: {p['name']}")
    
    print("===================\n")

def verify_problem_count(page, username: str, expected_count: int) -> None:
    """
    Navigate to Kattis profile and verify the problem count matches.
    """
    print("\n=== Verifying Problem Count on Kattis ===")
    
    # Go to the profile page
    page.goto(f"{KATTIS_BASE_URL}/users/{username}")
    page.wait_for_load_state("networkidle")
    
    # Try to find the solved problems count on the page
    html = page.content()
    soup = BeautifulSoup(html, 'html.parser')
    
    # Look for text like "207 problems solved"
    import re
    text = soup.get_text()
    
    # Try to find patterns like "X problems" or "Score: Y"
    problem_count_match = re.search(r'(\d+)\s+problems?\s+solved', text, re.I)
    score_match = re.search(r'Score[:\s]+(\d+\.?\d*)', text, re.I)
    
    if problem_count_match:
        kattis_count = int(problem_count_match.group(1))
        print(f"Kattis shows: {kattis_count} problems solved")
        print(f"We scraped:   {expected_count} problems")
        if kattis_count != expected_count:
            print(f"⚠️  MISMATCH: Missing {kattis_count - expected_count} problems!")
        else:
            print("✓ Problem count matches!")
    
    if score_match:
        kattis_score = float(score_match.group(1))
        print(f"\nKattis score: {kattis_score}")
    
    print("===================\n")

# -----------------------------------------------------------------------------
# Excel Parsing
# -----------------------------------------------------------------------------

def parse_excel_todo(excel_path: str) -> dict:
    """
    Parse the Grind.xlsx file to extract TODO items and daily tasks.
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    solved_from_excel = set()
    daily_tasks = {}
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        date_cell = row[0]
        date_val = date_cell.value
        
        # Process problem columns (starting from column G = index 6)
        for cell in row[6:]:
            if not cell.value:
                continue
            
            url = str(cell.value)
            if '/problems/' not in url:
                continue
            
            problem_id = url.split('/problems/')[-1].split('?')[0].split('/')[0]
            
            # Check if cell is marked as solved (green fill)
            is_solved = False
            fill = cell.fill
            if fill and fill.fill_type == 'solid' and fill.fgColor:
                # Theme 9 is typically green
                if fill.fgColor.type == 'theme' and fill.fgColor.theme == 9:
                    is_solved = True
            
            if is_solved:
                solved_from_excel.add(problem_id)
            
            # Track daily tasks
            if date_val and isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
                if date_str not in daily_tasks:
                    daily_tasks[date_str] = []
                daily_tasks[date_str].append({
                    'id': problem_id,
                    'url': f"{KATTIS_BASE_URL}/problems/{problem_id}",
                    'solved': is_solved
                })
    
    # Build TODO list from unsolved daily tasks
    todos = []
    for date_str in sorted(daily_tasks.keys()):
        for task in daily_tasks[date_str]:
            if not task['solved']:
                todos.append({
                    'date': date_str,
                    'id': task['id'],
                    'url': task['url']
                })
    
    return {
        'todos': todos,
        'solved_from_excel': list(solved_from_excel),
        'daily_tasks': daily_tasks
    }

# -----------------------------------------------------------------------------
# Local Solutions
# -----------------------------------------------------------------------------

def find_local_solutions(solutions_dir: str) -> set[str]:
    """
    Find all solution files in the solutions directory.
    Returns a set of problem IDs (filename stems).
    """
    solutions = set()
    
    if not os.path.exists(solutions_dir):
        return solutions
    
    for filename in os.listdir(solutions_dir):
        filepath = Path(filename)
        if filepath.suffix.lower() in SOLUTION_EXTENSIONS:
            solutions.add(filepath.stem)
    
    return solutions

def track_solution_dates(solutions_dir: str, tracking_file: str = 'solution_dates.xlsx') -> dict:
    """
    Track solution files and their modification dates in an Excel file.
    Returns a dict mapping problem_id -> date_string.
    """
    from openpyxl import load_workbook, Workbook
    from datetime import datetime
    
    # Load existing tracking file or create new one
    if os.path.exists(tracking_file):
        wb = load_workbook(tracking_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Solution Dates"
        ws['A1'] = 'Problem'
        ws['B1'] = 'Date'
    
    # Build dict of existing entries
    existing_entries = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value:
            existing_entries[row[0].value] = row[1].value
    
    # Scan solution directory
    if not os.path.exists(solutions_dir):
        wb.save(tracking_file)
        return existing_entries
    
    new_entries = []
    for filename in os.listdir(solutions_dir):
        filepath = Path(solutions_dir) / filename
        if filepath.suffix.lower() in SOLUTION_EXTENSIONS:
            problem_id = filepath.stem
            
            # Skip if already tracked
            if problem_id in existing_entries:
                continue
            
            # Get modification date
            mod_time = os.path.getmtime(filepath)
            mod_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')
            
            new_entries.append((problem_id, mod_date))
            existing_entries[problem_id] = mod_date
    
    # Append new entries to Excel
    if new_entries:
        print(f"  Adding {len(new_entries)} new solution date entries")
        for problem_id, date in sorted(new_entries, key=lambda x: x[1]):
            ws.append([problem_id, date])
        wb.save(tracking_file)
    
    return existing_entries

# -----------------------------------------------------------------------------
# Statistics
# -----------------------------------------------------------------------------

def track_score_over_time(total_score: float, tracking_file: str = 'scores.xlsx') -> dict:
    """
    Track total score over time in an Excel file.
    Creates the file if it doesn't exist, adds today's score.
    Returns a dict mapping date_string -> score for plotting.
    """
    from openpyxl import load_workbook, Workbook
    from datetime import datetime
    
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Load existing tracking file or create new one
    if os.path.exists(tracking_file):
        wb = load_workbook(tracking_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Score History"
        ws['A1'] = 'Date'
        ws['B1'] = 'Score'
        
        # Add initial data point for 2026-01-24 if creating new file
        ws.append(['2026-01-24', 596.4])
        print(f"  Created new score tracking file with baseline: 2026-01-24 -> 596.4")
    
    # Build dict of existing entries
    score_history = {}
    row_to_update = None
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value:
            date_str = row[0].value
            if isinstance(date_str, datetime):
                date_str = date_str.strftime('%Y-%m-%d')
            score_history[date_str] = row[1].value
            
            # Check if today's row exists
            if date_str == today:
                row_to_update = idx
    
    # Update or add today's score
    if row_to_update:
        ws.cell(row=row_to_update, column=2, value=total_score)
        print(f"  Updated score for {today}: {total_score}")
    else:
        ws.append([today, total_score])
        score_history[today] = total_score
        print(f"  Added new score entry for {today}: {total_score}")
    
    wb.save(tracking_file)
    return score_history

def categorize_difficulty(difficulty: float) -> str:
    """Map difficulty score to a category."""
    if difficulty < 2.8:
        return 'Easy (<2.8)'
    elif difficulty <= 5.5:
        return 'Medium (2.8-5.5)'
    else:
        return 'Hard (≥5.6)'

def generate_stats(
    username: str,
    solved_problems: list[dict],
    actual_total_score: float,
    local_solutions: set[str],
    solution_dates: dict,
    score_history: dict, 
    excel_data: Optional[dict] = None
) -> dict:
    """
    Generate comprehensive statistics from scraped data.
    """
    excel_data = excel_data or {}
    
    calculated_score = sum(p['difficulty'] for p in solved_problems)
    solved_ids = {p['id'] for p in solved_problems}
    
    # Compare with local solution files
    missing_solutions = sorted(solved_ids - local_solutions)
    extra_solutions = sorted(local_solutions - solved_ids)
    
    # Difficulty breakdown
    difficulty_buckets = defaultdict(int)
    for p in solved_problems:
        bucket = categorize_difficulty(p['difficulty'])
        difficulty_buckets[bucket] += 1
    
    # Build daily activity from solution dates
    daily_activity = defaultdict(int)
    for problem_id, date in solution_dates.items():
        daily_activity[date] += 1
    
    return {
        'generated_at': datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z'),
        'username': username,
        'total_score': round(actual_total_score, 1),
        'calculated_score': round(calculated_score, 1),
        'problems_solved': len(solved_problems),
        'problems': solved_problems,
        'missing_solutions': missing_solutions,
        'extra_solutions': extra_solutions,
        'difficulty_breakdown': dict(difficulty_buckets),
        'daily_activity': dict(daily_activity),
        'score_history': dict(score_history),  # Add this line
        'todos': excel_data.get('todos', []),
        'daily_tasks': excel_data.get('daily_tasks', {}),
        'target_score': 5000,
        'target_date': '2025-05-07'
    }

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='Scrape Kattis solved problems and generate statistics'
    )
    parser.add_argument('--username', required=True, help='Kattis username')
    parser.add_argument('--password', required=True, help='Kattis password')
    parser.add_argument('--solutions-dir', default='solutions',
                        help='Directory containing solution files')
    parser.add_argument('--excel-file', help='Path to Grind.xlsx TODO file')
    parser.add_argument('--output', default='docs/data.json',
                        help='Output JSON file path')
    parser.add_argument('--headed', action='store_true',
                        help='Run browser in headed mode (visible)')
    parser.add_argument('--debug', action='store_true',
                        help='Save debug HTML files')
    
    args = parser.parse_args()
    
    print(f"=== Kattis Stats Scraper ===\n")
    
    # Scrape Kattis
    solved_problems, actual_score = scrape_kattis(
        args.username,
        args.password,
        headless=not args.headed,
        debug=args.debug
    )
    
    # Find local solutions and track dates
    print(f"\nScanning solutions directory: {args.solutions_dir}")
    local_solutions = find_local_solutions(args.solutions_dir)
    print(f"  Found {len(local_solutions)} local solution files")

    print(f"\nTracking solution dates...")
    solution_dates = track_solution_dates(args.solutions_dir)
    print(f"  Tracking {len(solution_dates)} solutions with dates")

    print(f"\nTracking score over time...")
    score_history = track_score_over_time(actual_score)
    print(f"  Score history contains {len(score_history)} data points")
    
    # Parse Excel if provided
    excel_data = {}
    if args.excel_file and os.path.exists(args.excel_file):
        print(f"\nParsing Excel: {args.excel_file}")
        excel_data = parse_excel_todo(args.excel_file)
        print(f"  Found {len(excel_data.get('todos', []))} TODO items")
    
    # Generate stats
    # Generate stats
    # Generate stats
    stats = generate_stats(
        args.username, 
        solved_problems, 
        actual_score, 
        local_solutions, 
        solution_dates, 
        score_history, 
        excel_data
    )
    # Print missing solutions if any
    if stats['missing_solutions']:
        print(f"\n=== Missing Solutions (files without local record) ===")
        for solution_id in stats['missing_solutions']:
            print(f"  - {solution_id}")
        print()
    
    # Print extra solutions if any
    if stats['extra_solutions']:
        print(f"\n=== Extra Solutions (files without Kattis record) ===")
        for solution_id in stats['extra_solutions']:
            print(f"  - {solution_id}")
        print()

    
    # Write output
    output_dir = os.path.dirname(args.output)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    
    with open(args.output, 'w') as f:
        json.dump(stats, f, indent=2)
    
    # Print summary
    print(f"\n=== Summary ===")
    print(f"  Total Score:       {stats['total_score']}")
    print(f"  Problems Solved:   {stats['problems_solved']}")
    print(f"  Missing Solutions: {len(stats['missing_solutions'])}")
    print(f"  Extra Solutions:   {len(stats['extra_solutions'])}")
    print(f"\nOutput written to: {args.output}")


if __name__ == '__main__':
    main()